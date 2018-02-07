import socket
import datetime
import time
import sys
import os
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import binascii

IDS=[]
PIDS=[]
PID_LIST=[]
MODES=['01','02','03','09']
ELM_PROMPT = '>'
root_dir = os.getcwd()
minID=500
maxID=2048
minPID=0
maxPID=255
"""CAN-IDS to Loop over"""
for i in range(minID,maxID):
    i=(hex(i))[2:]
    i=format(i, '>02')
    IDS.append(i)
"""PIDS to Loop over"""
for i in range(minPID,maxPID):
    i=(hex(i))[2:]
    i=format(i, '>02')
    PIDS.append(i)

print '\nTool_SW_Vers: '
version=os.path.basename(__file__)
print version
print 'developed by Sebastian Kienitz'

#LOGGING-----------------------------------
file_name='INIT Log File_' + version+time.strftime("%Y_%m_%d_%Hh_%Mm")
te = open(file_name + '.txt','w')  # File where you need to keep the logs
print 'Log File: ' + file_name + '_' + time.strftime("%Y_%m_%d_%Hh_%Mm") + '.txt'

#------FUNCTIONS--------------------------------------------------------------------
def byte_to_binary(n):
    return ''.join(str((n & (1 << i)) and 1) for i in reversed(range(8)))

def hex_to_binary(h):
    return ''.join(byte_to_binary(ord(b)) for b in binascii.unhexlify(h))

def send_cmd(cmd):
    data=''
    cmd += "\r" # terminate
    s.send(cmd)
    i=0
    while True:
        data=data+s.recv(64)
        i=i+1
        print i
        if (data.endswith(ELM_PROMPT) or len(data)>128 or i>10):
            te.write(data)
            print data
            if len(data)>128:
                s.send("\r")
            break
    # remove the prompt character
    data = data[:-1]
    # splits into lines while removing empty lines and trailing spaces
    data = data.replace('\r','')
    data = data [len(cmd)-1:]
    """time.sleep(1)"""
    return data

"""INIT DONGLE-------------------------------------"""
s = socket.socket()
host = '192.168.0.10' # needs to be in quote
port = 35000
s.connect((host, port))
send_cmd("\r") 
send_cmd("ATSP0")
send_cmd("ATD")
protocol=send_cmd("ATDPN")[1:]
te.write('Protocol: ' + protocol+'\n')
te.write(send_cmd("ATI")+'\n')
send_cmd("0100")
te.write('Headers:')
te.write(send_cmd("ATH0")+'\n')
print send_cmd("ATAT2")
print send_cmd('ATSH' + IDS[len(IDS)-1])

#--------------------------MAIN---------------------------------------------------------------------------------------------------------------
print ('Battery Voltage:' + send_cmd('ATRV'))
te.write('Battery Voltage:' + send_cmd('ATRV')+'\n')
VIN=send_cmd('0902')
te.write('VIN ' + VIN+'\n')
if not (os.path.isfile(root_dir + '/' + 'PID.xlsx')):
    #VEHICLE-PRESCAN----------------------------------------------------
    print ('Learning valid PIDs ..................')
    for ID in IDS:
        print ID
        response =send_cmd('ATSH' + ID)
        print 'ATSH' + ID
        print response
        time.sleep(2)
        response=send_cmd("0100")
        if not('NO DAT' in response) and ('41' in response[:4]):
            for MODE in MODES:
                for PID in PIDS:
                    response=send_cmd(MODE + PID)
                    if not('NO DAT' in response) and ('4'+ MODE[1]in response[:4]):
                        print 'VALID OBD RESPONSE FOUND'
                        PID_LIST.append([ID,MODE,PID,response])

    te.write(str(PID_LIST))
    print 'No PID-LIST stored, saving list into PID.xlsx ..................'
    wb = Workbook()
    ws = wb.active
    print str(PID_LIST)
    for PID in PID_LIST:
        print 'SAVING LINE TO EXCEL: '
        print PID[0] + '-' + PID[1] + '-' + PID[2] + '-' + PID[3]
        ws.append([PID[0], PID[1], PID[2],PID[3]])
    wb.save('PID.xlsx')
else:
    print 'PID-LIST FOUND! ..................'
    wb=load_workbook(root_dir + '/' + 'PID.xlsx')
    first_sheet=wb.get_sheet_names()[0]
    ws=wb.get_sheet_by_name(first_sheet)
    for row in ws.iter_rows(min_row=1,max_col=3):
        line=[]
        for cell in row:
            if not(str(cell.value)=='None'):
                line.append(str(cell.value))
        PID_LIST.append(line)
"""delete last row cause it contains nothing"""
PID_LIST=PID_LIST[0:len(PID_LIST)-6]
te.write(str(PID_LIST))
te.close()

#LOGGING-----------------------------------
file_name='MEASUREMENT Log File_' + version+time.strftime("%Y_%m_%d_%Hh_%Mm")
te = open(file_name + '.txt','w')  # File where you need to keep the logs
print 'MEASUREMENT Log File: ' + file_name + '_' + time.strftime("%Y_%m_%d_%Hh_%Mm") + '.txt'

print 'Start Measurement of learned PIDs into Excel ..................'
wb = Workbook()
ws = wb.active
xls_lines=0
measurement_number=0
while True: 
    for PID in PID_LIST:
        send_cmd("ATSH" + PID[0])
        response=send_cmd(PID[1]+PID[2])
        ws.append([datetime.datetime.now(),PID[0], PID[1], PID[2],response])
        xls_lines=xls_lines+1
        if xls_lines>1000:
            wb.save(file_name + '_' + str(measurement_number) +'.xlsx')
            measurement_number=measurement_number+1
            wb = Workbook()
            ws = wb.active
            xls_lines=0
te.close()
